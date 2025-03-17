import os
import openpyxl


def save_to_excel_2(save_path, filename, columns, data):
    """
    保存数据到 Excel 文件，并确保文件夹存在。

    参数：
    - save_path (str): 目标文件夹路径
    - filename (str): Excel 文件名（包含 .xlsx 扩展名）
    - columns (list): Excel 表头（第一行）
    - data (list of lists): 需要写入的数据（从第二行开始）

    返回：
    - str: 保存的 Excel 文件路径
    """

    # 确保目标文件夹存在
    os.makedirs(save_path, exist_ok=True)

    # 创建 Excel 文件完整路径
    file_path = os.path.join(save_path, filename)

    # 创建 Excel 工作簿和工作表
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    # 写入列名（第一行）
    ws.append(columns)

    # 写入数据（从第二行开始）
    for row in data:
        ws.append(row)

    # 保存 Excel 文件
    wb.save(file_path)
    print(f"Excel 文件已保存至: {file_path}")

    return file_path  # 返回文件路径，方便调用时使用


def save_to_excel(save_path, filename, columns, data1, data2):
    """
    创建 Excel 文件，并在其中写入两个工作表的数据。

    参数：
    - save_path (str): Excel 存储的文件夹路径
    - filename (str): Excel 文件名（包含 .xlsx 扩展名）
    - columns (list): Excel 表头（第一行）
    - data1 (list of lists): 第一个工作表的数据
    - data2 (list of lists): 第二个工作表的数据

    返回：
    - str: 保存的 Excel 文件路径
    """

    # 确保目标文件夹存在
    os.makedirs(save_path, exist_ok=True)

    # 创建 Excel 文件完整路径
    file_path = os.path.join(save_path, filename + ".xlsx")

    # 创建 Excel 工作簿
    wb = openpyxl.Workbook()

    # 创建第一个工作表
    ws1 = wb.active  # 默认创建的第一个工作表
    ws1.title = "Gmean"
    ws1.append(columns)  # 写入表头
    for row in data1:
        ws1.append(row)  # 写入数据

    # 创建第二个工作表
    ws2 = wb.create_sheet(title="MAUC")
    ws2.append(columns)  # 写入表头
    for row in data2:
        ws2.append(row)  # 写入数据

    # 保存 Excel 文件
    wb.save(file_path)
    print(f"Excel 文件已保存至: {file_path}")

    return file_path  # 返回文件路径，方便调用时使用
